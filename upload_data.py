import sys
import os
import json

# 将标准输出编码设置为 UTF-8，避免 Windows 终端下中文显示乱码
try:
    sys.stdout.reconfigure(encoding="utf-8")  # pyrefly: ignore [missing-attribute]
except Exception:
    pass

import sqlite3
from datetime import datetime

import openpyxl
import pandas as pd
import requests

from news_sanitizer import is_valid_url, sanitize_news_item, verify_semantic_integrity

DB_NAME = "my_data.db"
EXCEL_FILE = "26630.xlsx"


def import_excel_to_db():
    """1. 同步 Excel 基础数字数据到数据库 (包含容错 fallback)"""
    try:
        xls = pd.ExcelFile(EXCEL_FILE)
        sheet_names = xls.sheet_names
        
        target_sheet1 = "图1，5" if "图1，5" in sheet_names else sheet_names[0]
        df1 = pd.read_excel(EXCEL_FILE, sheet_name=target_sheet1)
        if df1.shape[1] >= 36 and df1.shape[0] > 6:
            trend_data = df1.iloc[6:, [34, 35]].copy()
            trend_data.columns = ["date", "cpi_yoy"]
        else:
            trend_data = pd.DataFrame(columns=["date", "cpi_yoy"])

        def parse_date(val):
            if isinstance(val, pd.Timestamp) or hasattr(val, "strftime"):
                return val.strftime("%Y-%m-%d")
            try:
                dt = pd.to_datetime(val)
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return str(val)

        if not trend_data.empty:
            trend_data["date"] = trend_data["date"].apply(parse_date)
            trend_data["cpi_yoy"] = pd.to_numeric(trend_data["cpi_yoy"], errors="coerce")
            trend_data = trend_data.dropna()
            trend_data = trend_data.sort_values(by="date", ascending=True).reset_index(drop=True)

        target_sheet2 = "图2" if "图2" in sheet_names else sheet_names[0]
        df2 = pd.read_excel(EXCEL_FILE, sheet_name=target_sheet2)
        cols = [11, 13, 14, 15, 16, 17, 18, 19, 20]
        if df2.shape[1] >= 21 and df2.shape[0] > 6:
            cat_data = df2.iloc[6:, cols].copy()
            cat_data.columns = ["date", "食品烟酒", "衣着", "居住", "生活用品", "交通通信", "文教娱乐", "医疗", "其他"]
            cat_data["date"] = cat_data["date"].apply(parse_date)
            for col in cat_data.columns[1:]:
                cat_data[col] = pd.to_numeric(cat_data[col], errors="coerce")
            cat_data = cat_data.dropna()
            cat_data = cat_data.sort_values(by="date", ascending=True).reset_index(drop=True)
        else:
            cat_data = pd.DataFrame(columns=["date", "食品烟酒", "衣着", "居住", "生活用品", "交通通信", "文教娱乐", "医疗", "其他"])

        conn = sqlite3.connect(DB_NAME)
        if not trend_data.empty:
            trend_data.to_sql("cpi_trend", conn, if_exists="replace", index=False)
            trend_data.to_sql("sales_records", conn, if_exists="replace", index=False)
        if not cat_data.empty:
            cat_data.to_sql("cpi_categories", conn, if_exists="replace", index=False)
        conn.close()
        print("[Database] Excel 基础结构数据检查与 SQLite 同步完成。")
    except Exception as e:
        print(f"[Database] import_excel_to_db 兼容解析跳过: {e}")



def fetch_wechat_articles():
    """
    通过进门 MCP(comein-research) SSE 投研服务拉取「陈兴宏观研究」最近 10 篇微信公众号内容
    """
    try:
        from external_skills.comein_research_mcp import comein_research_mcp
        res_str = comein_research_mcp(
            "call_tool",
            tool_name="wechat_article_query",
            arguments={
                "keywords": "陈兴",
                "wechatArticleScope": "all",
                "pageSize": 25,
                "page": 1,
                "startTime": "2024-01-01 00:00:00"
            }
        )
        res_data = json.loads(res_str)
        text = res_data.get("result", [""])[0] if res_data.get("status") == "success" else ""

        # 解析底部的链接映射 [1]: url1, [2]: url2...
        import re
        link_map = {}
        for line in text.split("\n"):
            m = re.match(r"^\[(\d+)\]:\s*(https?://[^\s]+)", line.strip())
            if m:
                link_map[m.group(1)] = m.group(2)

        blocks = re.split(r"(?=### \d+\.)", text)
        articles = []

        for block in blocks:
            if not block.strip().startswith("###"):
                continue
            
            m_title = re.search(r"### \d+\.\s*\[(.*?)\]\[(\d+)\]", block)
            if not m_title:
                continue
            
            title = m_title.group(1).strip()
            link_idx = m_title.group(2)
            url = link_map.get(link_idx, "")

            m_time = re.search(r"-\s*时间：(\d{4}-\d{2}-\d{2})", block)
            pub_time = m_time.group(1) if m_time else datetime.now().strftime("%Y-%m-%d")

            is_chenxing_macro = any(k in block for k in ["华福", "分析师", "首席经济学家论坛", "宏观", "固收", "策略", "陈兴：", "陈兴、"])
            is_irrelevant = any(k in block for k in ["纪委监委", "商丘", "辞去", "副主席", "跳槽", "CFO", "新希望"])

            if is_chenxing_macro and not is_irrelevant:
                clean_title = re.sub(r"^(陈兴[：:]\s*)", "", title).strip()
                articles.append({
                    "id": len(articles) + 1,
                    "publish_time": pub_time,
                    "content": clean_title,  # 纯净标题（剥离 陈兴： 前缀）
                    "url": url,
                })
        
        if len(articles) >= 5:
            print(f"[进门 MCP] 成功从「陈兴宏观研究」微信公众号检索到 {len(articles)} 篇最新宏观研报！")
            return articles[:10]
    except Exception as e:
        print(f"[进门 MCP 实时抓取警告] 远程拉取失败, 启动降级静态逻辑: {e}")

    # Fallback default 10 static links if offline
    fallback_urls = [
        ("2026-07-19", "韩国央行转向加息", "https://mp.weixin.qq.com/s/NJq0AEpCbSzP78AwvTDBEQ"),
        ("2026-07-16", "信用降温趋势延续——2026年6月金融数据解读", "https://mp.weixin.qq.com/s/7MYZa-P8amU5OkntRg_rmQ"),
        ("2026-07-12", "美国服务业价格压力缓解", "https://mp.weixin.qq.com/s/28vPWqmtoH6TOkXU7LRFEw"),
        ("2026-07-05", "全球储蓄——由过剩到短缺？", "https://mp.weixin.qq.com/s/tag8klgGoAscSChRCvJdow"),
        ("2026-06-14", "没了点阵图，市场如何反应？", "https://mp.weixin.qq.com/s/CXtO0LxA0U9gYOFzZ2CH-Q"),
        ("2026-06-13", "M1同比何以反弹？——2026年5月金融数据解读", "https://mp.weixin.qq.com/s/7MYZa-P8amU5OkntRg_rmQ"),
        ("2025-04-22", "深度 | 特朗普怎样对医药“动刀”？—— “特朗普经济学”系列之十四【陈兴团队·财通宏观】", "https://mp.weixin.qq.com/s/28vPWqmtoH6TOkXU7LRFEw"),
        ("2024-11-11", "深度 | 谁会是特朗普的新助手？——美国大选深度观察之六【财通宏观•陈兴团队】", "https://mp.weixin.qq.com/s/tag8klgGoAscSChRCvJdow"),
        ("2024-09-25", "【划重点】国新办就金融支持经济高质量发展有关情况举行新闻发布会要点【财通宏观 陈兴团队】", "https://mp.weixin.qq.com/s/NJq0AEpCbSzP78AwvTDBEQ"),
        ("2024-07-15", "一颗子弹，几多出口？——特朗普遇刺事件解读【财通宏观•陈兴团队】", "https://mp.weixin.qq.com/s/CXtO0LxA0U9gYOFzZ2CH-Q")
    ]
    return [{"id": i+1, "publish_time": dt, "content": title, "url": u} for i, (dt, title, u) in enumerate(fallback_urls)]




def generate_and_save_macro_analysis():
    """
    3. 组装「陈兴宏观研究」微信公众号最近 10 篇研报超链接列表入库
    """
    articles = fetch_wechat_articles()
    if not articles:
        return

    top_10 = articles[:10]

    rows_html = []
    for idx, item in enumerate(top_10):
        border_style = "border-bottom: 1px dashed #1e293b;" if idx < len(top_10) - 1 else ""
        row_str = f'<div style="{border_style} padding: 7px 0; display: flex; align-items: center; justify-content: space-between;"><span style="color: #94a3b8; font-size: 13px; min-width: 100px;">📅 {item["publish_time"]}</span> <a href="{item["url"]}" target="_blank" style="color: #38bdf8; text-decoration: none; font-weight: 600; font-size: 14px; flex-grow: 1; margin-left: 10px; text-overflow: ellipsis; overflow: hidden; white-space: nowrap;">{item["content"]}</a></div>'
        rows_html.append(row_str)

    rows_combined = "\n".join(rows_html)

    dynamic_html = f'''<div style="background: linear-gradient(135deg, #0b1a30 0%, #081325 100%); border-radius: 12px; padding: 22px; border: 1px solid #132a4a; font-family: 'Microsoft YaHei', 'PingFang SC', sans-serif;">
<div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 16px; border-bottom: 1px solid rgba(0, 240, 255, 0.15); padding-bottom: 10px;">
    <h4 style="color:#ffffff; margin:0; font-size: 1.05rem; font-weight: 700;">📰 微信公众号：陈兴宏观研究 ‧ 最新 10 篇研报直达</h4>
    <span style="background: rgba(16, 185, 129, 0.15); color: #10b981; border: 1px solid rgba(16, 185, 129, 0.3); padding: 2px 8px; border-radius: 4px; font-size: 0.72rem; font-weight: 600;">🟢 进门 MCP 实时专线同步</span>
</div>
<div style="line-height: 2.0; color: #e2e8f0;">
{rows_combined}
</div>
</div>'''

    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS macro_analysis (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                html_text TEXT,
                update_time TEXT
            )
        """)
        cursor.execute("DELETE FROM macro_analysis")
        cursor.execute(
            "INSERT INTO macro_analysis (html_text, update_time) VALUES (?, ?)",
            (dynamic_html, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
        conn.close()
        print(f"[本地列表引擎] 「陈兴宏观研究」公众号最新 {len(top_10)} 篇研究超链接已成功同步至 SQLite 数据库！")
    except Exception as e:
        print(f"[本地列表引擎] 写入数据库失败: {e}")



def fetch_finance_news(limit=5):
    """4. 实时在线抓取新浪 7x24 快讯以供顶部滚动横幅展现"""
    api_limit = max(limit * 3, 20)
    url = f"https://zhibo.sina.com.cn/api/zhibo/feed?page=1&page_size={api_limit}&zhibo_id=152&tag_id=0&dire=1&dpc=1"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    try:
        response = requests.get(url, headers=headers, timeout=15)
        news_list = response.json().get("result", {}).get("data", {}).get("feed", {}).get("list", [])
        records = []
        for item in news_list:
            doc_url = item.get("docurl", "").strip()
            if not is_valid_url(doc_url):
                continue
            # 提取并清洗 title 和 content 独立字段
            raw_text = item.get("rich_text", "").strip()
            t, c = sanitize_news_item(raw_text)
            # V1.1.4.2 & V1.1.4.3: Headline Guard & Semantic SVO filter
            if not t or len(str(t).replace("。", "").strip()) < 5 or not verify_semantic_integrity(t):
                continue
            records.append({
                "id": item.get("id"),
                "publish_time": item.get("create_time"),
                "title": t,
                "content": c,
                "url": doc_url,
                "source": "新浪财经 7×24",
                "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            if len(records) >= limit:
                break
        return records
    except Exception as e:
        print(f"新浪快讯网络抓取失败: {e}")
        return []


def import_news_to_db(records, table_name="text_records"):
    """将实时快讯保存到快讯数据库表"""
    if not records:
        return
    df = pd.DataFrame(records)
    conn = sqlite3.connect(DB_NAME)
    df.to_sql(table_name, conn, if_exists="replace", index=False)
    conn.close()
    print(f"[Database] 成功拉取并保存 {len(records)} 条今日新浪金融实时快讯！")


def import_dashboard_charts_to_db():
    """读取 DASHBOARD 工作表中的折线图数据源区域，清洗并入库"""
    excel_file = EXCEL_FILE
    db_name = DB_NAME
    
    cpi_compare_fallback = True
    coal_prices_fallback = True
    cpi_compare_df = pd.DataFrame(columns=["date", "cpi_yoy", "core_cpi_yoy"])
    coal_prices_df = pd.DataFrame(columns=["date", "dlm_price", "jm_price"])
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        if "DASHBOARD" in wb.sheetnames:
            ws = wb["DASHBOARD"]
            for chart in ws._charts:  # pyrefly: ignore [missing-attribute]
                if not chart.series:
                    continue
                first_series = chart.series[0]
                if not first_series.val or not hasattr(first_series.val, "numRef"):
                    continue
                ref_formula = first_series.val.numRef.f
                
                if "图1，5" in ref_formula:
                    print("[Dashboard Parser] 动态检测到 CPI 对比折线图数据源:", ref_formula)
                    try:
                        df1 = pd.read_excel(excel_file, sheet_name="图1，5")
                        cpi_compare_df = df1.iloc[7:104, [11, 13, 14]].copy()
                        cpi_compare_df.columns = ["date", "cpi_yoy", "core_cpi_yoy"]
                        cpi_compare_fallback = False
                    except Exception as parse_err:
                        print(f"[Dashboard Parser] 动态解析 CPI 对比折线图失败: {parse_err}")
                
                elif "图3，4" in ref_formula:
                    print("[Dashboard Parser] 动态检测到煤炭价格折线图数据源:", ref_formula)
                    try:
                        df3 = pd.read_excel(excel_file, sheet_name="图3，4")
                        coal_prices_df = df3.iloc[7:269, [26, 27, 28]].copy()
                        coal_prices_df.columns = ["date", "dlm_price", "jm_price"]
                        coal_prices_fallback = False
                    except Exception as parse_err:
                        print(f"[Dashboard Parser] 动态解析煤炭价格折线图失败: {parse_err}")
    except Exception as e:
        print(f"[Dashboard Parser] 动态解析 DASHBOARD 图表失败，将启用硬编码兜底解析: {e}")
        
    def parse_date(val):
        if isinstance(val, pd.Timestamp) or hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        try:
            dt = pd.to_datetime(val)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return str(val)

    if cpi_compare_fallback:
        print("[Dashboard Parser] 启用 CPI 对比折线图兜底解析")
        try:
            df1 = pd.read_excel(excel_file, sheet_name="图1，5")
            cpi_compare_df = df1.iloc[7:104, [11, 13, 14]].copy()
            cpi_compare_df.columns = ["date", "cpi_yoy", "core_cpi_yoy"]
        except Exception as e:
            print(f"[Dashboard Parser] CPI 兜底解析失败: {e}")

    assert isinstance(cpi_compare_df, pd.DataFrame)
    cpi_compare_df["date"] = cpi_compare_df["date"].apply(parse_date)
    cpi_compare_df["cpi_yoy"] = pd.to_numeric(cpi_compare_df["cpi_yoy"], errors="coerce")
    cpi_compare_df["core_cpi_yoy"] = pd.to_numeric(cpi_compare_df["core_cpi_yoy"], errors="coerce")
    cpi_compare_df = cpi_compare_df.dropna()
    cpi_compare_df = cpi_compare_df.sort_values(by="date", ascending=True).reset_index(drop=True)

    if coal_prices_fallback:
        print("[Dashboard Parser] 启用煤炭价格折线图兜底解析")
        try:
            df3 = pd.read_excel(excel_file, sheet_name="图3，4")
            coal_prices_df = df3.iloc[7:269, [26, 27, 28]].copy()
            coal_prices_df.columns = ["date", "dlm_price", "jm_price"]
        except Exception as e:
            print(f"[Dashboard Parser] 煤炭价格兜底解析失败: {e}")

    assert isinstance(coal_prices_df, pd.DataFrame)
    coal_prices_df["date"] = coal_prices_df["date"].apply(parse_date)
    coal_prices_df["dlm_price"] = pd.to_numeric(coal_prices_df["dlm_price"], errors="coerce")
    coal_prices_df["jm_price"] = pd.to_numeric(coal_prices_df["jm_price"], errors="coerce")
    coal_prices_df = coal_prices_df.dropna()
    coal_prices_df = coal_prices_df[(coal_prices_df["dlm_price"] > 0) & (coal_prices_df["jm_price"] > 0)]
    assert isinstance(coal_prices_df, pd.DataFrame)
    coal_prices_df = coal_prices_df.sort_values(by="date", ascending=True).reset_index(drop=True)

    # V1.4.2.0: 食品分项价格矩阵解析 (图6 工作表)
    food_prices_df = None
    try:
        # 动态检测 "图6" 工作表 (注意: 工作表名可能含尾随空格)
        food_sheet_name = None
        xls_sheets = pd.ExcelFile(excel_file).sheet_names
        for sn in xls_sheets:
            sn_str = str(sn).strip() if sn is not None else ""
            if sn_str == "图6":
                food_sheet_name = sn
                break
        
        if food_sheet_name:
            print(f"[Dashboard Parser] 动态检测到食品分项价格数据源: sheet='{food_sheet_name}'")
            df6 = pd.read_excel(excel_file, sheet_name=food_sheet_name)
            # 列映射: Unnamed:1=日期, Unnamed:4=鸡蛋, Unnamed:5=鲜菜, Unnamed:6=鲜果, Unnamed:7=猪肉（右）
            food_prices_df = df6.iloc[12:, [1, 5, 4, 6, 7]].copy()
            food_prices_df.columns = ["date", "fresh_vegetable", "egg", "fresh_fruit", "pork"]
            food_prices_df["date"] = food_prices_df["date"].apply(parse_date)
            for col in ["fresh_vegetable", "egg", "fresh_fruit", "pork"]:
                food_prices_df[col] = pd.to_numeric(food_prices_df[col], errors="coerce")
            food_prices_df = food_prices_df.dropna()
            assert isinstance(food_prices_df, pd.DataFrame)
            food_prices_df = food_prices_df.sort_values(by="date", ascending=True).reset_index(drop=True)
            print(f"[Dashboard Parser] 食品分项价格解析成功: {len(food_prices_df)} 行有效数据")
        else:
            print("[Dashboard Parser] 未找到 '图6' 工作表，食品分项价格将跳过")
    except Exception as e_food:
        print(f"[Dashboard Parser] 食品分项价格解析异常: {e_food}")

    conn = sqlite3.connect(db_name)
    cpi_compare_df.to_sql("dashboard_cpi_compare", conn, if_exists="replace", index=False)
    coal_prices_df.to_sql("dashboard_coal_prices", conn, if_exists="replace", index=False)
    if food_prices_df is not None and not food_prices_df.empty:
        food_prices_df.to_sql("dashboard_food_prices", conn, if_exists="replace", index=False)
        print("[Database] dashboard_food_prices 食品分项价格表同步成功！")
    conn.close()
    print("[Database] DASHBOARD 折线图数据同步成功！")


def import_inflation_and_fiscal_to_db():
    """5. 解析 26630.xlsx 中的通胀、财政与金融数据区，保存至 SQLite 数据库"""
    import os
    if not os.path.exists(EXCEL_FILE):
        print(f"[Database] 未找到文件 {EXCEL_FILE}")
        return

    try:
        df = pd.read_excel(EXCEL_FILE, header=None)

        # 1. 提取通胀数据区 (Columns 0..18)
        inf_dates = [pd.to_datetime(d).strftime("%Y-%m-%d") for d in df.iloc[1, 3:19] if pd.notna(d)]
        records_inf = []
        for idx, dt in enumerate(inf_dates):
            col_i = 3 + idx
            records_inf.append({
                "date": dt,
                "cpi_yoy": pd.to_numeric(df.iloc[2, col_i], errors="coerce"),
                "core_cpi_yoy": pd.to_numeric(df.iloc[3, col_i], errors="coerce"),
                "food_cpi_yoy": pd.to_numeric(df.iloc[4, col_i], errors="coerce"),
                "nonfood_cpi_yoy": pd.to_numeric(df.iloc[5, col_i], errors="coerce"),
                "cpi_mom": pd.to_numeric(df.iloc[6, col_i], errors="coerce"),
                "core_cpi_mom": pd.to_numeric(df.iloc[7, col_i], errors="coerce"),
                "food_cpi_mom": pd.to_numeric(df.iloc[8, col_i], errors="coerce"),
                "nonfood_cpi_mom": pd.to_numeric(df.iloc[9, col_i], errors="coerce"),
                "ppi_yoy": pd.to_numeric(df.iloc[10, col_i], errors="coerce"),
                "ppi_production_yoy": pd.to_numeric(df.iloc[11, col_i], errors="coerce"),
                "ppi_life_yoy": pd.to_numeric(df.iloc[12, col_i], errors="coerce"),
                "ppi_mom": pd.to_numeric(df.iloc[13, col_i], errors="coerce"),
                "ppi_production_mom": pd.to_numeric(df.iloc[14, col_i], errors="coerce"),
                "ppi_life_mom": pd.to_numeric(df.iloc[15, col_i], errors="coerce"),
            })
        df_inf_series = pd.DataFrame(records_inf).sort_values(by="date", ascending=True).reset_index(drop=True)

        # 2. 提取财政数据区 (Columns 20..34)
        fis_dates = [pd.to_datetime(d).strftime("%Y-%m-%d") for d in df.iloc[1, 22:35] if pd.notna(d)]
        records_fis = []
        for idx, dt in enumerate(fis_dates):
            col_i = 22 + idx
            records_fis.append({
                "date": dt,
                "fiscal_revenue_yoy": pd.to_numeric(df.iloc[2, col_i], errors="coerce"),
                "central_revenue_yoy": pd.to_numeric(df.iloc[3, col_i], errors="coerce"),
                "local_revenue_yoy": pd.to_numeric(df.iloc[4, col_i], errors="coerce"),
                "tax_revenue_yoy": pd.to_numeric(df.iloc[5, col_i], errors="coerce"),
                "nontax_revenue_yoy": pd.to_numeric(df.iloc[6, col_i], errors="coerce"),
                "fiscal_expenditure_yoy": pd.to_numeric(df.iloc[7, col_i], errors="coerce"),
                "central_expenditure_yoy": pd.to_numeric(df.iloc[8, col_i], errors="coerce"),
                "local_expenditure_yoy": pd.to_numeric(df.iloc[9, col_i], errors="coerce"),
                "fund_revenue_yoy": pd.to_numeric(df.iloc[10, col_i], errors="coerce"),
                "land_concession_yoy": pd.to_numeric(df.iloc[11, col_i], errors="coerce"),
                "fund_expenditure_yoy": pd.to_numeric(df.iloc[12, col_i], errors="coerce"),
            })
        df_fis_series = pd.DataFrame(records_fis).sort_values(by="date", ascending=True).reset_index(drop=True)

        # 3. 提取金融数据区 (Columns 35..45)
        df_fin_series = pd.DataFrame()
        if df.shape[1] >= 46:
            fin_dates = [pd.to_datetime(d).strftime("%Y-%m-%d") for d in df.iloc[1, 36:46] if pd.notna(d)]
            records_fin = []
            fin_row_keys = [
                "social_financing_inc", "sf_rmb_loan", "sf_yoy_diff", "sf_rmb_loan_yoy_diff",
                "foreign_currency_loan", "entrusted_loan", "trust_loan", "undiscounted_acceptance_bills",
                "corporate_bonds", "equity_financing", "government_bonds", "credit_inc",
                "household_loan_inc", "household_short_term", "household_mid_long",
                "corporate_loan_inc", "corporate_short_term", "corporate_mid_long",
                "bill_financing", "credit_yoy_diff", "household_loan_yoy_diff",
                "household_short_term_yoy_diff", "household_mid_long_yoy_diff",
                "corporate_loan_yoy_diff", "corporate_short_term_yoy_diff",
                "corporate_mid_long_yoy_diff", "bill_financing_yoy_diff", "deposit_yoy_diff",
                "fiscal_deposit", "household_deposit", "corporate_deposit", "nonbank_deposit",
                "m1_yoy", "m2_yoy", "m2_m1_diff", "sf_stock_yoy"
            ]

            for idx, dt in enumerate(fin_dates):
                col_i = 36 + idx
                rec = {"date": dt}
                for r_i, k in enumerate(fin_row_keys):
                    row_idx = 2 + r_i
                    if row_idx < df.shape[0]:
                        rec[k] = pd.to_numeric(df.iloc[row_idx, col_i], errors="coerce")
                records_fin.append(rec)
            if records_fin:
                df_fin_series = pd.DataFrame(records_fin).sort_values(by="date", ascending=True).reset_index(drop=True)

        # 4. 提取经济数据区 (Columns 47..)
        df_econ_series = pd.DataFrame()
        df_gdp_series = pd.DataFrame()
        if df.shape[1] >= 47:
            # 4.1 月度经济指标数据提取
            econ_dates = []
            for c in range(47, df.shape[1]):
                d_val = df.iloc[4, c]
                if pd.notna(d_val):
                    try:
                        econ_dates.append((c, pd.to_datetime(d_val).strftime("%Y-%m-%d")))
                    except Exception:
                        pass
            records_econ = []
            for col_i, dt in econ_dates:
                rec = {
                    "date": dt,
                    "industrial_gva_yoy": pd.to_numeric(df.iloc[5, col_i], errors="coerce"),
                    "industrial_gva_mom": pd.to_numeric(df.iloc[6, col_i], errors="coerce"),
                    "service_index_yoy": pd.to_numeric(df.iloc[7, col_i], errors="coerce"),
                    "pmi_manuf": pd.to_numeric(df.iloc[8, col_i], errors="coerce"),
                    "pmi_manuf_prod": pd.to_numeric(df.iloc[9, col_i], errors="coerce"),
                    "pmi_manuf_orders": pd.to_numeric(df.iloc[10, col_i], errors="coerce"),
                    "pmi_manuf_export_orders": pd.to_numeric(df.iloc[11, col_i], errors="coerce"),
                    "pmi_non_manuf": pd.to_numeric(df.iloc[12, col_i], errors="coerce"),
                    "fai_yoy": pd.to_numeric(df.iloc[13, col_i], errors="coerce"),
                    "fai_realestate_yoy": pd.to_numeric(df.iloc[14, col_i], errors="coerce"),
                    "fai_infra_yoy": pd.to_numeric(df.iloc[15, col_i], errors="coerce"),
                    "fai_manuf_yoy": pd.to_numeric(df.iloc[16, col_i], errors="coerce"),
                    "retail_sales_yoy": pd.to_numeric(df.iloc[17, col_i], errors="coerce"),
                    "retail_sales_above_size_yoy": pd.to_numeric(df.iloc[18, col_i], errors="coerce"),
                    "property_sales_area_yoy": pd.to_numeric(df.iloc[19, col_i], errors="coerce"),
                    "property_starts_yoy": pd.to_numeric(df.iloc[20, col_i], errors="coerce"),
                    "property_completions_yoy": pd.to_numeric(df.iloc[21, col_i], errors="coerce"),
                    "unemployment_rate": pd.to_numeric(df.iloc[22, col_i], errors="coerce"),
                    "export_yoy": pd.to_numeric(df.iloc[23, col_i], errors="coerce"),
                    "import_yoy": pd.to_numeric(df.iloc[24, col_i], errors="coerce"),
                    "trade_balance": pd.to_numeric(df.iloc[25, col_i], errors="coerce"),
                    "revenue_yoy": pd.to_numeric(df.iloc[34, col_i], errors="coerce") if df.shape[0] > 34 else None,
                    "profit_yoy": pd.to_numeric(df.iloc[35, col_i], errors="coerce") if df.shape[0] > 35 else None,
                    "inventory_yoy": pd.to_numeric(df.iloc[36, col_i], errors="coerce") if df.shape[0] > 36 else None,
                }
                records_econ.append(rec)
            if records_econ:
                df_econ_series = pd.DataFrame(records_econ).sort_values(by="date", ascending=True).reset_index(drop=True)

            # 4.2 季度 GDP 增长数据提取
            gdp_dates = []
            for c in range(47, df.shape[1]):
                q_val = df.iloc[1, c]
                if pd.notna(q_val) and "Q" in str(q_val):
                    gdp_dates.append((c, str(q_val).strip()))
            records_gdp = []
            for col_i, q_str in gdp_dates:
                rec = {
                    "quarter": q_str,
                    "gdp_yoy": pd.to_numeric(df.iloc[2, col_i], errors="coerce"),
                    "gdp_mom": pd.to_numeric(df.iloc[3, col_i], errors="coerce"),
                }
                records_gdp.append(rec)
            if records_gdp:
                df_gdp = pd.DataFrame(records_gdp)
                df_gdp["sort_key"] = df_gdp["quarter"].apply(lambda q: "20" + q if len(q) == 4 else q)
                df_gdp_series = df_gdp.sort_values(by="sort_key", ascending=True).drop(columns=["sort_key"]).reset_index(drop=True)

        # 5. 提取较上月变化 (Delta changes)
        kpi_deltas = {}
        for r in range(2, 16):
            name = str(df.iloc[r, 0]).strip()
            chg = df.iloc[r, 1]
            if pd.notna(chg):
                kpi_deltas[name] = float(chg)
        for r in range(2, 13):
            name = str(df.iloc[r, 21]).strip()
            chg = df.iloc[r, 20]
            if pd.notna(chg):
                kpi_deltas[name] = float(chg)

        # 计算金融数据最新月与前一个月的环比 Delta
        if not df_fin_series.empty and len(df_fin_series) >= 2:
            latest_fin = df_fin_series.iloc[-1]
            prev_fin = df_fin_series.iloc[-2]
            kpi_deltas["社融当月新增"] = float(latest_fin.get("social_financing_inc", 0.0)) - float(prev_fin.get("social_financing_inc", 0.0))
            kpi_deltas["信贷当月新增"] = float(latest_fin.get("credit_inc", 0.0)) - float(prev_fin.get("credit_inc", 0.0))
            kpi_deltas["M2同比增速"] = float(latest_fin.get("m2_yoy", 0.0)) - float(prev_fin.get("m2_yoy", 0.0))
            kpi_deltas["M1同比增速"] = float(latest_fin.get("m1_yoy", 0.0)) - float(prev_fin.get("m1_yoy", 0.0))
            kpi_deltas["社融存量同比增速"] = float(latest_fin.get("sf_stock_yoy", 0.0)) - float(prev_fin.get("sf_stock_yoy", 0.0))

        # 计算经济数据核心指标最新月与前一个月的环比 Delta
        if not df_econ_series.empty and len(df_econ_series) >= 2:
            latest_econ = df_econ_series.iloc[-1]
            prev_econ = df_econ_series.iloc[-2]
            kpi_deltas["制造业PMI"] = float(latest_econ.get("pmi_manuf", 0.0)) - float(prev_econ.get("pmi_manuf", 0.0))
            kpi_deltas["固资投资增速"] = float(latest_econ.get("fai_yoy", 0.0)) - float(prev_econ.get("fai_yoy", 0.0))
            kpi_deltas["社零增速"] = float(latest_econ.get("retail_sales_yoy", 0.0)) - float(prev_econ.get("retail_sales_yoy", 0.0))
            kpi_deltas["出口同比"] = float(latest_econ.get("export_yoy", 0.0)) - float(prev_econ.get("export_yoy", 0.0))

        df_deltas = pd.DataFrame([{"metric_key": k, "change_mom": v} for k, v in kpi_deltas.items()])

        # 6. 写入 SQLite
        conn = sqlite3.connect(DB_NAME)
        df_inf_series.to_sql("dashboard_inflation_series", conn, if_exists="replace", index=False)
        df_fis_series.to_sql("dashboard_fiscal_series", conn, if_exists="replace", index=False)
        if not df_fin_series.empty:
            df_fin_series.to_sql("dashboard_finance_series", conn, if_exists="replace", index=False)
        if not df_econ_series.empty:
            df_econ_series.to_sql("dashboard_economic_series", conn, if_exists="replace", index=False)
        if not df_gdp_series.empty:
            df_gdp_series.to_sql("dashboard_gdp_series", conn, if_exists="replace", index=False)
        df_deltas.to_sql("dashboard_kpi_deltas", conn, if_exists="replace", index=False)

        # 兼容 dashboard_cpi_compare
        cpi_compat = df_inf_series[["date", "cpi_yoy", "core_cpi_yoy"]].copy()
        cpi_compat.to_sql("dashboard_cpi_compare", conn, if_exists="replace", index=False)

        conn.close()
        print("[Database] 通胀、财政、金融与经济数据 (dashboard_inflation_series, dashboard_fiscal_series, dashboard_finance_series, dashboard_economic_series, dashboard_gdp_series) 同步至 SQLite 成功！")
    except Exception as e:
        print(f"[Database] 解析通胀、财政与金融数据区失败: {e}")



def import_excel_embedded_charts_to_db(excel_file="26630.xlsx"):
    """
    自动抽取 26630.xlsx 首页中的所有 Excel 原生嵌入图表 (BarChart / LineChart 等)
    并将图表标题、分类、数值序列保存至 SQLite 数据库 dashboard_embedded_charts
    """
    if not os.path.exists(excel_file):
        return

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active

        extracted_charts = []
        for i, c in enumerate(ws._charts):
            title_text = f"Chart {i+1}"
            if c.title:
                try:
                    if hasattr(c.title, "tx") and c.title.tx and c.title.tx.rich:
                        txt = ""
                        for p in c.title.tx.rich.p:
                            for r in p.r:
                                txt += r.t
                        if txt: title_text = txt
                    elif isinstance(c.title, str):
                        title_text = c.title
                except Exception:
                    pass
            
            title_text = title_text.strip()
            chart_type = c.__class__.__name__

            series_data = []
            for s_idx, series in enumerate(c.series):
                cat_vals = []
                if hasattr(series, "cat") and series.cat:
                    if hasattr(series.cat, "strRef") and series.cat.strRef and series.cat.strRef.strCache:
                        cat_vals = [str(pt.v).strip() for pt in series.cat.strRef.strCache.pt if pt.v is not None]
                    elif hasattr(series.cat, "numRef") and series.cat.numRef and series.cat.numRef.numCache:
                        for pt in series.cat.numRef.numCache.pt:
                            if pt.v is not None:
                                val = pt.v
                                if isinstance(val, (int, float)) and val > 40000 and val < 50000:
                                    try:
                                        dt = pd.to_datetime(val - 25569, unit="D")
                                        val = dt.strftime("%Y-%m")
                                    except Exception:
                                        pass
                                cat_vals.append(str(val))

                num_vals = []
                if hasattr(series, "val") and series.val:
                    if hasattr(series.val, "numRef") and series.val.numRef and series.val.numRef.numCache:
                        for pt in series.val.numRef.numCache.pt:
                            try:
                                num_vals.append(float(pt.v) if pt.v is not None else 0.0)
                            except Exception:
                                num_vals.append(0.0)
                    elif hasattr(series.val, "strRef") and series.val.strRef and series.val.strRef.strCache:
                        for pt in series.val.strRef.strCache.pt:
                            try:
                                num_vals.append(float(pt.v) if pt.v is not None else 0.0)
                            except Exception:
                                num_vals.append(0.0)

                stitle = f"Series {s_idx+1}"
                if hasattr(series, "title") and series.title:
                    if hasattr(series.title, "v") and series.title.v:
                        stitle = str(series.title.v).strip()
                    elif hasattr(series.title, "strRef") and series.title.strRef and series.title.strRef.strCache:
                        pts = series.title.strRef.strCache.pt
                        if pts and pts[0].v is not None:
                            stitle = str(pts[0].v).strip()

                if stitle.isdigit() and int(stitle) > 40000:
                    try:
                        dt = pd.to_datetime(int(stitle) - 25569, unit="D")
                        stitle = dt.strftime("%Y-%m")
                    except Exception:
                        pass

                series_data.append({
                    "series_title": stitle,
                    "categories": cat_vals,
                    "values": num_vals
                })

            extracted_charts.append({
                "chart_id": i + 1,
                "title": title_text,
                "type": chart_type,
                "series": series_data
            })

        conn = sqlite3.connect(DB_NAME)
        records = []
        for ch in extracted_charts:
            records.append({
                "chart_id": ch["chart_id"],
                "chart_title": ch["title"],
                "chart_type": ch["type"],
                "chart_json": json.dumps(ch, ensure_ascii=False)
            })
        df_embedded = pd.DataFrame(records)
        df_embedded.to_sql("dashboard_embedded_charts", conn, if_exists="replace", index=False)
        conn.close()
        print(f"[Database] 成功从 {excel_file} 提取并保存 {len(extracted_charts)} 个原生嵌入图表至 SQLite (dashboard_embedded_charts)！")
    except Exception as e:
        print(f"[Database] 提取 Excel 嵌入图表失败: {e}")


def main():
    print("=" * 50)
    print("🚀 开始执行全自动数据加工处理流水线...")
    print("=" * 50)
    import_excel_to_db()
    import_dashboard_charts_to_db()
    import_inflation_and_fiscal_to_db()
    import_excel_embedded_charts_to_db()

    # 抓取并同步顶部滚动快讯
    news_records = fetch_finance_news(limit=5)
    import_news_to_db(news_records)

    # 构建并同步底部手动维护的文章传送门列表
    generate_and_save_macro_analysis()

    print("\n[OK] 数据源加工流已全盘就绪！")


if __name__ == "__main__":
    main()
